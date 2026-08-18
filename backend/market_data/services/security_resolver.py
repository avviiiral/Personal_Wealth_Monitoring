import re
from pathlib import Path

from django.conf import settings


class SecurityResolver:
    """
    Resolve Indian securities to Yahoo Finance symbols.

    Resolution priority:
        1. Security Master Excel ISIN mapping
        2. Existing hard-coded ISIN mapping
        3. Explicit Yahoo/exchange symbol
        4. Security Master Excel name mapping
        5. Existing hard-coded name mapping
        6. Generic symbol + exchange suffix

    The Security Master is optional. If it is unavailable, the existing
    resolver behaviour remains available as a fallback.
    """

    NSE_SUFFIX = ".NS"
    BSE_SUFFIX = ".BO"

    SECURITY_MASTER_FILENAME = "security_master.xlsx"

    # ==========================================================
    # Existing fallback ISIN -> Yahoo Finance symbol mapping
    # ==========================================================

    ISIN_TO_YAHOO = {
        "INE021A01026": "ASIANPAINT.NS",
        "INE397D01024": "BHARTIARTL.NS",
        "INE860A01027": "HCLTECH.NS",
        "INE040A01034": "HDFCBANK.NS",
        "INE030A01027": "HINDUNILVR.NS",
        "INE090A01021": "ICICIBANK.NS",
        "INE154A01025": "ITC.NS",
        "INE009A01021": "INFY.NS",
        "INE018A01030": "LT.NS",
        "INE101A01026": "M&M.NS",
        "INE585B01010": "MARUTI.NS",
        "INE002A01018": "RELIANCE.NS",
        "INE062A01020": "SBIN.NS",
        "INE044A01036": "SUNPHARMA.NS",
        "INE467B01029": "TCS.NS",
        "INE155A01022": "TATAMOTORS.NS",
    }

    # ==========================================================
    # Existing fallback asset-name mapping
    # ==========================================================

    NAME_TO_YAHOO = {
        "ASIAN PAINTS LTD": "ASIANPAINT.NS",
        "ASIAN PAINTS": "ASIANPAINT.NS",

        "BHARTI AIRTEL LTD": "BHARTIARTL.NS",
        "BHARTI AIRTEL": "BHARTIARTL.NS",

        "HCL TECHNOLOGIES": "HCLTECH.NS",
        "HCL TECHNOLOGIES LTD": "HCLTECH.NS",
        "HCL TECHNOLOGIES LIMITED": "HCLTECH.NS",

        "HDFC BANK LTD": "HDFCBANK.NS",
        "HDFC BANK LIMITED": "HDFCBANK.NS",
        "HDFC BANK": "HDFCBANK.NS",

        "HINDUSTAN UNILEVER LTD": "HINDUNILVR.NS",
        "HINDUSTAN UNILEVER LIMITED": "HINDUNILVR.NS",
        "HINDUSTAN UNILEVER": "HINDUNILVR.NS",

        "ICICI BANK LTD": "ICICIBANK.NS",
        "ICICI BANK LIMITED": "ICICIBANK.NS",
        "ICICI BANK": "ICICIBANK.NS",

        "ITC LTD": "ITC.NS",
        "ITC LIMITED": "ITC.NS",
        "ITC": "ITC.NS",

        "INFOSYS LTD": "INFY.NS",
        "INFOSYS LIMITED": "INFY.NS",
        "INFOSYS": "INFY.NS",

        "LARSEN & TOUBRO LTD": "LT.NS",
        "LARSEN & TOUBRO LIMITED": "LT.NS",
        "LARSEN AND TOUBRO LTD": "LT.NS",
        "LARSEN AND TOUBRO LIMITED": "LT.NS",

        "MAHINDRA & MAHINDRA LTD": "M&M.NS",
        "MAHINDRA & MAHINDRA LIMITED": "M&M.NS",
        "MAHINDRA AND MAHINDRA LTD": "M&M.NS",

        "MARUTI SUZUKI INDIA LTD": "MARUTI.NS",
        "MARUTI SUZUKI INDIA LIMITED": "MARUTI.NS",
        "MARUTI SUZUKI": "MARUTI.NS",

        "RELIANCE INDUSTRIES": "RELIANCE.NS",
        "RELIANCE INDUSTRIES LTD": "RELIANCE.NS",
        "RELIANCE INDUSTRIES LIMITED": "RELIANCE.NS",

        "STATE BANK OF INDIA": "SBIN.NS",
        "STATE BANK OF INDIA LTD": "SBIN.NS",
        "STATE BANK OF INDIA LIMITED": "SBIN.NS",

        "SUN PHARMACEUTICAL INDUSTRIES LTD": "SUNPHARMA.NS",
        "SUN PHARMACEUTICAL INDUSTRIES LIMITED": "SUNPHARMA.NS",
        "SUN PHARMA": "SUNPHARMA.NS",

        "TATA CONSULTANCY SERVICES LTD": "TCS.NS",
        "TATA CONSULTANCY SERVICES LIMITED": "TCS.NS",
        "TCS": "TCS.NS",

        "TATA MOTORS LTD": "TATAMOTORS.NS",
        "TATA MOTORS LIMITED": "TATAMOTORS.NS",
        "TATA MOTORS": "TATAMOTORS.NS",
    }

    _security_master_loaded = False
    _security_master_isin = {}
    _security_master_name = {}

    @staticmethod
    def clean_symbol(symbol):
        if not symbol:
            return ""

        value = str(symbol).strip().upper()

        value = re.sub(
            r"\s*\(NSE\)\s*$",
            "",
            value,
        )

        value = re.sub(
            r"\s*\(BSE\)\s*$",
            "",
            value,
        )

        value = value.replace(" ", "")

        return value

    @staticmethod
    def clean_isin(isin):
        if not isin:
            return ""

        return str(isin).strip().upper()

    @staticmethod
    def clean_name(name):
        if not name:
            return ""

        value = str(name).strip().upper()

        value = re.sub(
            r"\s+",
            " ",
            value,
        )

        return value

    @classmethod
    def _security_master_path(cls):
        """
        Locate security_master.xlsx.

        Preferred location:
            backend/data/security_master.xlsx

        settings.BASE_DIR points to the backend directory in the
        current Django project.
        """

        base_dir = Path(settings.BASE_DIR)

        candidates = [
            base_dir / "data" / cls.SECURITY_MASTER_FILENAME,
            base_dir.parent / "data" / cls.SECURITY_MASTER_FILENAME,
        ]

        for path in candidates:
            if path.exists():
                return path

        return None

    @classmethod
    def _load_security_master(cls):
        """
        Load Security Master Excel once per process.

        The workbook is optional. If it is missing or unreadable,
        the resolver continues using its existing mappings.
        """

        if cls._security_master_loaded:
            return

        cls._security_master_loaded = True

        path = cls._security_master_path()

        if path is None:
            return

        try:
            import openpyxl

            workbook = openpyxl.load_workbook(
                path,
                read_only=True,
                data_only=True,
            )

            worksheet = workbook.active

            headers = next(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=1,
                    values_only=True,
                ),
                (),
            )

            normalized_headers = {}

            for index, header in enumerate(headers):
                if header is None:
                    continue

                normalized_headers[
                    str(header).strip().upper()
                ] = index

            isin_index = normalized_headers.get("ISIN")
            yahoo_index = normalized_headers.get("YAHOO SYMBOL")
            name_index = normalized_headers.get("SECURITY NAME")

            if isin_index is None:
                workbook.close()
                return

            for row in worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ):
                if not row:
                    continue

                isin_value = (
                    row[isin_index]
                    if isin_index < len(row)
                    else None
                )

                yahoo_value = (
                    row[yahoo_index]
                    if yahoo_index is not None
                    and yahoo_index < len(row)
                    else None
                )

                name_value = (
                    row[name_index]
                    if name_index is not None
                    and name_index < len(row)
                    else None
                )

                cleaned_isin = cls.clean_isin(isin_value)
                cleaned_yahoo = cls.clean_symbol(yahoo_value)
                cleaned_name = cls.clean_name(name_value)

                if (
                    cleaned_isin
                    and cleaned_yahoo
                ):
                    cls._security_master_isin[
                        cleaned_isin
                    ] = cleaned_yahoo

                if (
                    cleaned_name
                    and cleaned_yahoo
                ):
                    cls._security_master_name[
                        cleaned_name
                    ] = cleaned_yahoo

            workbook.close()

        except Exception:
            # Security Master must not break the existing price
            # resolution system if the Excel file is unavailable,
            # malformed, or openpyxl is not installed.
            cls._security_master_isin = {}
            cls._security_master_name = {}

    @classmethod
    def reload_security_master(cls):
        """
        Force the Security Master to be loaded again.

        Useful after security_master.xlsx has been edited while
        the Django process is still running.
        """

        cls._security_master_loaded = False
        cls._security_master_isin = {}
        cls._security_master_name = {}

        cls._load_security_master()

    @classmethod
    def resolve_from_isin(cls, isin):
        """
        Resolve Yahoo symbol directly from ISIN.

        Security Master is checked first, followed by the existing
        hard-coded mapping.
        """

        cleaned_isin = cls.clean_isin(isin)

        if not cleaned_isin:
            return None

        cls._load_security_master()

        yahoo_from_master = cls._security_master_isin.get(
            cleaned_isin
        )

        if yahoo_from_master:
            return yahoo_from_master

        return cls.ISIN_TO_YAHOO.get(
            cleaned_isin
        )

    @classmethod
    def resolve_from_name(cls, name):
        """
        Resolve Yahoo symbol from a known asset name.

        Security Master is checked first, followed by the existing
        hard-coded mapping.
        """

        cleaned_name = cls.clean_name(name)

        if not cleaned_name:
            return None

        cls._load_security_master()

        yahoo_from_master = cls._security_master_name.get(
            cleaned_name
        )

        if yahoo_from_master:
            return yahoo_from_master

        return cls.NAME_TO_YAHOO.get(
            cleaned_name
        )

    @classmethod
    def resolve_yahoo_symbol(
        cls,
        symbol=None,
        exchange="NSE",
        isin=None,
        name=None,
    ):
        """
        Resolve the Yahoo Finance symbol.

        Priority:

            1. ISIN through Security Master
            2. Existing ISIN mapping
            3. Explicit symbol
            4. Security Master name mapping
            5. Existing name mapping
            6. Generic symbol + exchange suffix
        """

        yahoo_from_isin = cls.resolve_from_isin(
            isin
        )

        if yahoo_from_isin:
            return yahoo_from_isin

        cleaned_symbol = cls.clean_symbol(
            symbol
        )

        if cleaned_symbol:

            if cleaned_symbol.endswith(
                cls.NSE_SUFFIX
            ):
                return cleaned_symbol

            if cleaned_symbol.endswith(
                cls.BSE_SUFFIX
            ):
                return cleaned_symbol

        yahoo_from_name = cls.resolve_from_name(
            name
        )

        if yahoo_from_name:
            return yahoo_from_name

        if not cleaned_symbol:
            raise ValueError(
                "Unable to resolve security symbol. "
                "ISIN, symbol and asset name are all missing "
                "or unmapped."
            )

        exchange = (
            exchange or "NSE"
        ).strip().upper()

        if exchange == "BSE":
            return (
                f"{cleaned_symbol}"
                f"{cls.BSE_SUFFIX}"
            )

        return (
            f"{cleaned_symbol}"
            f"{cls.NSE_SUFFIX}"
        )

    @classmethod
    def candidate_symbols(
        cls,
        symbol=None,
        exchange=None,
        isin=None,
        name=None,
    ):
        """
        Return possible Yahoo symbols.

        Security Master and ISIN mappings are preferred.
        """

        candidates = []

        yahoo_from_isin = cls.resolve_from_isin(
            isin
        )

        if yahoo_from_isin:
            candidates.append(
                yahoo_from_isin
            )

        yahoo_from_name = cls.resolve_from_name(
            name
        )

        if (
            yahoo_from_name
            and yahoo_from_name not in candidates
        ):
            candidates.append(
                yahoo_from_name
            )

        cleaned = cls.clean_symbol(
            symbol
        )

        if cleaned:

            if cleaned.endswith(
                cls.NSE_SUFFIX
            ):

                if cleaned not in candidates:
                    candidates.append(cleaned)

            elif cleaned.endswith(
                cls.BSE_SUFFIX
            ):

                if cleaned not in candidates:
                    candidates.append(cleaned)

            else:

                exchange = (
                    exchange or "NSE"
                ).strip().upper()

                if exchange == "BSE":
                    candidate = (
                        f"{cleaned}"
                        f"{cls.BSE_SUFFIX}"
                    )
                else:
                    candidate = (
                        f"{cleaned}"
                        f"{cls.NSE_SUFFIX}"
                    )

                if candidate not in candidates:
                    candidates.append(candidate)

        return candidates